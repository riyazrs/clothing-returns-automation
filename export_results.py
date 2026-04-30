"""Export Returns_Queue.xlsx to JSON for GitHub Pages display."""
import openpyxl
import json
import os
from datetime import datetime

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "data", "Returns_Queue.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "docs", "results.json")

def export():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d")
                record[headers[i]] = val
        if record.get("Order_ID"):
            rows.append(record)

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w") as f:
        json.dump({"timestamp": datetime.utcnow().isoformat() + "Z", "results": rows}, f, indent=2, default=str)

    print(f"Exported {len(rows)} rows to {OUTPUT_FILE}")

if __name__ == "__main__":
    export()
